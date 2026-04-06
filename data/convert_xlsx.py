"""
God's Eye Sachsen — Excel → JSON Konverter
Konvertiert Gemeindeverzeichnis_Sachsen.xlsx in gemeinden.json

Verwendung:
    pip install openpyxl
    python convert_xlsx.py "Pfad/zur/Gemeindeverzeichnis_Sachsen.xlsx"

Die Ausgabe wird als gemeinden.json im gleichen Verzeichnis gespeichert.
"""

import json
import sys
import os

try:
    from openpyxl import load_workbook
except ImportError:
    print("Fehler: openpyxl nicht installiert. Bitte ausführen:")
    print("  pip install openpyxl")
    sys.exit(1)


# Spalten-Mapping (0-basiert) — angepasst an Gemeindeverzeichnis_Sachsen.xlsx
COLUMN_MAP = {
    0: 'id',
    1: 'schlnr',
    2: 'nr_krs',
    3: 'land',
    4: 'bundesland',
    5: 'landkreis',
    6: 'gemeinde',
    7: 'status_typ',
    8: 'plz',
    9: 'ortsteile',
    10: 'psf_plz',
    11: 'psf_ort',
    12: 'psf_strasse',
    13: 'okz',
    14: 'telefon',
    15: 'fax',
    16: 'email',
    17: 'homepage',
    18: 'bm_name',
    19: 'bm_titel',
    20: 'bm_anrede',
    21: 'bm_status',
    22: 'bm_sonstiges',
    23: 'regionalplan',
    24: 'fnp',
    25: 'bplan',
    26: 'energiekonzept',
    27: 'konflikte',
    28: 'wichtig',
    29: 'bauamt_kontakt',
    30: 'todo',
    31: 'wiedervorlage',
    32: 'gis_zuordnung',
    33: 'bearb_stand',
    34: 'sonstiges',
    35: 'extra'
}


def convert(xlsx_path):
    """Konvertiert die Excel-Datei in JSON"""
    print(f"Lade: {xlsx_path}")
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    gemeinden = []
    header_found = False

    for row in ws.iter_rows(min_row=1, values_only=True):
        # Header-Zeile überspringen (erste Zeile mit Daten erkennen)
        if not header_found:
            # Prüfe ob diese Zeile wie ein Header aussieht
            first_val = str(row[0] or '').strip()
            if first_val.isdigit() or (len(first_val) > 0 and first_val[0].isdigit()):
                header_found = True
            else:
                continue

        # Leere Zeilen überspringen
        if not row[0] and not row[6]:
            continue

        gemeinde = {}
        for col_idx, field_name in COLUMN_MAP.items():
            if col_idx < len(row):
                val = row[col_idx]
                # None → leerer String
                if val is None:
                    val = ''
                # Zahlen als String (für SCHLNR etc.)
                elif isinstance(val, (int, float)):
                    val = str(int(val)) if float(val) == int(val) else str(val)
                else:
                    val = str(val).strip()
                gemeinde[field_name] = val

        # Nur Einträge mit Gemeindename
        if gemeinde.get('gemeinde'):
            gemeinden.append(gemeinde)

    wb.close()

    # JSON speichern
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'gemeinden.json')
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(gemeinden, f, ensure_ascii=False, indent=2)

    print(f"Fertig: {len(gemeinden)} Gemeinden → {output_path}")
    return gemeinden


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Verwendung: python convert_xlsx.py <Pfad/zur/Excel-Datei>")
        print()
        print("Beispiel:")
        print('  python convert_xlsx.py "C:\\Users\\lnavarro\\...\\Gemeindeindex_für_claude.xlsx"')
        sys.exit(1)

    convert(sys.argv[1])
